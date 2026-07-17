#!/usr/bin/env python3
"""Generate all Day 028 notebooks: exercises 1-5, project, solution."""
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DAY_DIR = ROOT / "02_automation" / "day_028"

_cid = 0


def cid():
    global _cid
    _cid += 1
    return f"c{_cid:04d}"


def md(source: str) -> dict:
    return {"cell_type": "markdown", "id": cid(), "metadata": {}, "source": source}


def code(source: str) -> dict:
    return {
        "cell_type": "code",
        "id": cid(),
        "metadata": {},
        "outputs": [],
        "execution_count": None,
        "source": source,
    }


def nb(cells: list) -> dict:
    return {
        "nbformat": 4,
        "nbformat_minor": 5,
        "metadata": {
            "kernelspec": {
                "display_name": "ai-course",
                "language": "python",
                "name": "ai-course",
            },
            "language_info": {"name": "python", "version": "3.11.0"},
        },
        "cells": cells,
    }


def write_nb(path: Path, cells: list):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(nb(cells), indent=1), encoding="utf-8")
    print(f"  wrote {path.relative_to(ROOT)}")


# ---------------------------------------------------------------------------
# Implementations
# ---------------------------------------------------------------------------

READ_ROWS_IMPL = """\
def read_sheet_rows(path: str, sheet_name: str | None = None) -> list[dict]:
    wb = load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]"""

WRITE_ROWS_IMPL = """\
def write_sheet_rows(
    path: str,
    sheet_name: str,
    headers: list[str],
    rows: list[dict],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(path)"""

APPEND_ROW_IMPL = """\
def append_sheet_row(path: str, sheet_name: str, row: dict) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    headers = [
        ws.cell(row=1, column=c).value
        for c in range(1, ws.max_column + 1)
    ]
    ws.append([row.get(h) for h in headers])
    wb.save(path)"""

BOLD_HEADER_IMPL = """\
def bold_header_row(path: str, sheet_name: str) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    wb.save(path)"""

AI_ANALYZE_IMPL = """\
def ai_analyze_sheet(path: str, question: str, model: str = "llama3.2") -> str:
    rows = read_sheet_rows(path)
    data_str = json.dumps(rows[:20], indent=2)
    response = ollama.chat(
        model=model,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a data analyst. Answer questions about tabular data "
                    "concisely and precisely."
                ),
            },
            {
                "role": "user",
                "content": f"Data:\\n{data_str[:3000]}\\n\\nQuestion: {question}",
            },
        ],
    )
    return response["message"]["content"]"""

ANALYST_IMPL = """\
class SpreadsheetAnalyst:
    def __init__(self, model: str = "llama3.2"):
        self.model = model

    def analyze(self, rows: list[dict], question: str) -> str:
        data_str = json.dumps(rows[:20], indent=2)
        response = ollama.chat(
            model=self.model,
            messages=[
                {
                    "role": "system",
                    "content": "You are a data analyst. Answer concisely.",
                },
                {
                    "role": "user",
                    "content": f"Data:\\n{data_str[:3000]}\\n\\nQuestion: {question}",
                },
            ],
        )
        return response["message"]["content"]

    def write_summary(
        self, input_path: str, analysis: str, output_path: str
    ) -> None:
        rows = read_sheet_rows(input_path)
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Data"
        if rows:
            headers = list(rows[0].keys())
            ws_data.append(headers)
            for row in rows:
                ws_data.append([row.get(h) for h in headers])
            for cell in ws_data[1]:
                cell.font = Font(bold=True)
        ws_sum = wb.create_sheet("Summary")
        ws_sum.append(["AI Analysis"])
        ws_sum.append([analysis])
        wb.save(output_path)

    def run(
        self, input_path: str, question: str, output_path: str
    ) -> dict:
        rows = read_sheet_rows(input_path)
        analysis = self.analyze(rows, question)
        self.write_summary(input_path, analysis, output_path)
        return {
            "rows_analyzed": len(rows),
            "analysis": analysis,
            "output": output_path,
        }\
"""

ALL_IMPLS = "\n\n\n".join([
    READ_ROWS_IMPL,
    WRITE_ROWS_IMPL,
    APPEND_ROW_IMPL,
    BOLD_HEADER_IMPL,
    AI_ANALYZE_IMPL,
])


# ---------------------------------------------------------------------------
# Exercise 01 — read_sheet_rows
# ---------------------------------------------------------------------------

def ex01():
    global _cid; _cid = 0
    return [
        md(
            "# Day 028 — Exercise 1: read_sheet_rows\n\n"
            "**What you'll build:** `read_sheet_rows(path, sheet_name=None) -> list[dict]` — "
            "reads an XLSX file and returns all data rows as a list of dicts, with the first "
            "row used as column headers.\n\n"
            "**Why it matters:** This is the XLSX equivalent of `csv.DictReader` from Day 21 — "
            "same `list[dict]` output shape, same header-from-first-row convention. "
            "A consistent read interface means the rest of your pipeline doesn't care "
            "whether data came from a CSV or a spreadsheet."
        ),
        code("from openpyxl import load_workbook, Workbook"),
        md("## Your Implementation"),
        code(
            "def read_sheet_rows(\n"
            "    path: str,\n"
            "    sheet_name: str | None = None,\n"
            ") -> list[dict]:\n"
            '    """\n'
            "    Read an XLSX file and return data rows as list[dict].\n\n"
            "    Args:\n"
            "        path:       Path to the .xlsx file.\n"
            "        sheet_name: Worksheet name to read. Defaults to the active sheet.\n\n"
            "    Returns:\n"
            "        List of dicts — first row as keys, remaining rows as values.\n"
            "        Empty list if the sheet has no rows.\n"
            '    """\n'
            "    # TODO: wb = load_workbook(path)\n"
            "    # TODO: ws = wb[sheet_name] if sheet_name else wb.active\n"
            "    # TODO: rows = list(ws.iter_rows(values_only=True))\n"
            "    # TODO: if not rows: return []\n"
            "    # TODO: headers = [str(h) for h in rows[0]]\n"
            "    # TODO: return [dict(zip(headers, row)) for row in rows[1:]]\n"
            "    pass"
        ),
        md("## Check Your Work"),
        code(
            "import tempfile, os\n"
            "\n"
            "\n"
            "def _run_checks():\n"
            "    total = 5\n"
            "    passed = 0\n"
            "\n"
            "    tmp = tempfile.mktemp(suffix='.xlsx')\n"
            "    try:\n"
            "        # Setup: build a test XLSX\n"
            "        wb = Workbook()\n"
            "        ws = wb.active\n"
            "        ws.title = 'Data'\n"
            "        ws.append(['Name', 'Score', 'Status'])\n"
            "        ws.append(['Alice', 92, 'Pass'])\n"
            "        ws.append(['Bob', 74, 'Pass'])\n"
            "        ws.append(['Carol', 55, 'Fail'])\n"
            "        wb.save(tmp)\n"
            "\n"
            "        # Check 1: defined\n"
            "        try:\n"
            "            assert 'read_sheet_rows' in globals()\n"
            "            passed += 1; print('\\u2705 Check 1: read_sheet_rows defined')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 1: {e}')\n"
            "            print(f'\\nScore: {passed}/{total}')\n"
            "            return\n"
            "\n"
            "        result = None\n"
            "\n"
            "        # Check 2: returns a list\n"
            "        try:\n"
            "            result = read_sheet_rows(tmp)\n"
            "            assert isinstance(result, list), \\\n"
            "                f'expected list, got {type(result)}'\n"
            "            passed += 1; print(f'\\u2705 Check 2: returns a list ({len(result)} items)')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 2: {e}')\n"
            "\n"
            "        # Check 3: correct row count (header excluded)\n"
            "        try:\n"
            "            assert result is not None\n"
            "            assert len(result) == 3, \\\n"
            "                f'expected 3 data rows, got {len(result)}'\n"
            "            passed += 1; print(f'\\u2705 Check 3: {len(result)} data rows (header excluded)')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 3: {e}')\n"
            "\n"
            "        # Check 4: rows are dicts with correct keys\n"
            "        try:\n"
            "            assert result is not None and len(result) > 0\n"
            "            row = result[0]\n"
            "            assert isinstance(row, dict), \\\n"
            "                f'expected dict, got {type(row)}'\n"
            "            assert set(row.keys()) == {'Name', 'Score', 'Status'}, \\\n"
            "                f'keys wrong: {list(row.keys())}'\n"
            "            passed += 1; print('\\u2705 Check 4: rows are dicts with correct keys')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 4: {e}')\n"
            "\n"
            "        # Check 5: first row values correct\n"
            "        try:\n"
            "            assert result is not None and len(result) > 0\n"
            "            row = result[0]\n"
            "            assert row['Name'] == 'Alice', \\\n"
            "                f\"Name wrong: {row['Name']!r}\"\n"
            "            assert row['Score'] == 92, \\\n"
            "                f\"Score wrong: {row['Score']!r}\"\n"
            "            assert row['Status'] == 'Pass', \\\n"
            "                f\"Status wrong: {row['Status']!r}\"\n"
            "            passed += 1; print('\\u2705 Check 5: first row values correct')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 5: {e}')\n"
            "\n"
            "    finally:\n"
            "        try:\n"
            "            os.unlink(tmp)\n"
            "        except Exception:\n"
            "            pass\n"
            "\n"
            "    if passed == total:\n"
            "        print('\\U0001f389 Exercise complete!')\n"
            "    print(f'\\nScore: {passed}/{total}')\n"
            "\n"
            "\n"
            "_run_checks()"
        ),
        md(
            "## Solution\n\n"
            "<details>\n"
            "<summary>Click to reveal</summary>\n\n"
            "```python\n"
            + READ_ROWS_IMPL + "\n"
            "```\n\n"
            "</details>"
        ),
    ]


# ---------------------------------------------------------------------------
# Exercise 02 — write_sheet_rows
# ---------------------------------------------------------------------------

def ex02():
    global _cid; _cid = 100
    return [
        md(
            "# Day 028 — Exercise 2: write_sheet_rows\n\n"
            "**What you'll build:** `write_sheet_rows(path, sheet_name, headers, rows)` — "
            "creates a new XLSX file, writes a named worksheet with the given headers as "
            "the first row, and writes each row dict as a subsequent row in column order.\n\n"
            "**Why it matters:** This is the XLSX equivalent of `csv.DictWriter` from Day 21. "
            "Standardising on `list[dict]` for both read and write means a CSV-to-XLSX "
            "converter is just `write_sheet_rows(out, 'Sheet1', headers, read_csv(in))`."
        ),
        code("from openpyxl import load_workbook, Workbook"),
        md("## Your Implementation"),
        code(
            "def write_sheet_rows(\n"
            "    path: str,\n"
            "    sheet_name: str,\n"
            "    headers: list[str],\n"
            "    rows: list[dict],\n"
            ") -> None:\n"
            '    """\n'
            "    Create a new XLSX file with one sheet containing the given rows.\n\n"
            "    Args:\n"
            "        path:       Output file path (created or overwritten).\n"
            "        sheet_name: Name for the active worksheet.\n"
            "        headers:    Column names — written as the first row.\n"
            "        rows:       List of dicts — each written as a data row in header order.\n"
            '    """\n'
            "    # TODO: wb = Workbook(). ws = wb.active. ws.title = sheet_name\n"
            "    # TODO: ws.append(headers)\n"
            "    # TODO: for row in rows: ws.append([row.get(h) for h in headers])\n"
            "    # TODO: wb.save(path)\n"
            "    pass"
        ),
        md("## Check Your Work"),
        code(
            "import tempfile, os\n"
            "\n"
            "\n"
            "def _run_checks():\n"
            "    total = 5\n"
            "    passed = 0\n"
            "\n"
            "    tmp = tempfile.mktemp(suffix='.xlsx')\n"
            "    HEADERS = ['Product', 'Q1', 'Q2']\n"
            "    ROWS = [\n"
            "        {'Product': 'Alpha', 'Q1': 1200, 'Q2': 1350},\n"
            "        {'Product': 'Beta',  'Q1':  870, 'Q2':  920},\n"
            "        {'Product': 'Gamma', 'Q1': 2100, 'Q2': 1980},\n"
            "    ]\n"
            "\n"
            "    # Check 1: defined\n"
            "    try:\n"
            "        assert 'write_sheet_rows' in globals()\n"
            "        passed += 1; print('\\u2705 Check 1: write_sheet_rows defined')\n"
            "    except Exception as e:\n"
            "        print(f'\\u274c Check 1: {e}')\n"
            "        print(f'\\nScore: {passed}/{total}')\n"
            "        return\n"
            "\n"
            "    try:\n"
            "        # Check 2: creates the file\n"
            "        try:\n"
            "            write_sheet_rows(tmp, 'Sales', HEADERS, ROWS)\n"
            "            assert os.path.exists(tmp), 'file not created'\n"
            "            passed += 1; print('\\u2705 Check 2: file created at path')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 2: {e}')\n"
            "\n"
            "        # Check 3: sheet name is correct\n"
            "        try:\n"
            "            wb = load_workbook(tmp)\n"
            "            assert 'Sales' in wb.sheetnames, \\\n"
            "                f\"sheet 'Sales' not found: {wb.sheetnames}\"\n"
            "            passed += 1; print(\"\\u2705 Check 3: sheet named 'Sales' exists\")\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 3: {e}')\n"
            "\n"
            "        # Check 4: header row correct\n"
            "        try:\n"
            "            wb = load_workbook(tmp)\n"
            "            ws = wb['Sales']\n"
            "            header_row = [ws.cell(row=1, column=c).value\n"
            "                          for c in range(1, 4)]\n"
            "            assert header_row == HEADERS, \\\n"
            "                f'header wrong: {header_row}'\n"
            "            passed += 1; print(f'\\u2705 Check 4: header row is {header_row}')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 4: {e}')\n"
            "\n"
            "        # Check 5: data rows correct count and values\n"
            "        try:\n"
            "            wb = load_workbook(tmp)\n"
            "            ws = wb['Sales']\n"
            "            data_rows = list(ws.iter_rows(\n"
            "                min_row=2, values_only=True\n"
            "            ))\n"
            "            assert len(data_rows) == 3, \\\n"
            "                f'expected 3 data rows, got {len(data_rows)}'\n"
            "            assert data_rows[0][0] == 'Alpha', \\\n"
            "                f\"first product wrong: {data_rows[0][0]!r}\"\n"
            "            assert data_rows[0][1] == 1200, \\\n"
            "                f\"Q1 wrong: {data_rows[0][1]!r}\"\n"
            "            passed += 1; print(f'\\u2705 Check 5: {len(data_rows)} data rows, values correct')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 5: {e}')\n"
            "\n"
            "    finally:\n"
            "        try:\n"
            "            os.unlink(tmp)\n"
            "        except Exception:\n"
            "            pass\n"
            "\n"
            "    if passed == total:\n"
            "        print('\\U0001f389 Exercise complete!')\n"
            "    print(f'\\nScore: {passed}/{total}')\n"
            "\n"
            "\n"
            "_run_checks()"
        ),
        md(
            "## Solution\n\n"
            "<details>\n"
            "<summary>Click to reveal</summary>\n\n"
            "```python\n"
            + WRITE_ROWS_IMPL + "\n"
            "```\n\n"
            "</details>"
        ),
    ]


# ---------------------------------------------------------------------------
# Exercise 03 — append_sheet_row
# ---------------------------------------------------------------------------

def ex03():
    global _cid; _cid = 200
    return [
        md(
            "# Day 028 — Exercise 3: append_sheet_row\n\n"
            "**What you'll build:** `append_sheet_row(path, sheet_name, row)` — "
            "opens an existing XLSX file, reads the header row to determine column order, "
            "and appends the given dict as a new row at the bottom.\n\n"
            "**Why it matters:** Automation scripts typically run repeatedly and add new "
            "rows to a running log or tracker. `append_sheet_row` handles this pattern: "
            "load, read headers, append, save — preserving all existing data."
        ),
        code("from openpyxl import load_workbook, Workbook"),
        md("## Your Implementation"),
        code(
            "def append_sheet_row(path: str, sheet_name: str, row: dict) -> None:\n"
            '    """\n'
            "    Append a row dict to an existing XLSX worksheet.\n\n"
            "    Args:\n"
            "        path:       Path to the existing .xlsx file.\n"
            "        sheet_name: Worksheet to append to.\n"
            "        row:        Dict of {column_name: value} — missing keys become None.\n"
            '    """\n'
            "    # TODO: wb = load_workbook(path). ws = wb[sheet_name]\n"
            "    # TODO: headers = [ws.cell(row=1, column=c).value\n"
            "    #                   for c in range(1, ws.max_column + 1)]\n"
            "    # TODO: ws.append([row.get(h) for h in headers])\n"
            "    # TODO: wb.save(path)\n"
            "    pass"
        ),
        md("## Check Your Work"),
        code(
            "import tempfile, os\n"
            "\n"
            "\n"
            "def _run_checks():\n"
            "    total = 5\n"
            "    passed = 0\n"
            "\n"
            "    tmp = tempfile.mktemp(suffix='.xlsx')\n"
            "    try:\n"
            "        # Setup: XLSX with 2 existing data rows\n"
            "        wb = Workbook()\n"
            "        ws = wb.active\n"
            "        ws.title = 'Tasks'\n"
            "        ws.append(['Task', 'Status', 'Points'])\n"
            "        ws.append(['Task A', 'Done', 10])\n"
            "        ws.append(['Task B', 'Done', 20])\n"
            "        wb.save(tmp)\n"
            "\n"
            "        # Check 1: defined\n"
            "        try:\n"
            "            assert 'append_sheet_row' in globals()\n"
            "            passed += 1; print('\\u2705 Check 1: append_sheet_row defined')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 1: {e}')\n"
            "            print(f'\\nScore: {passed}/{total}')\n"
            "            return\n"
            "\n"
            "        # Check 2: row count increases by 1\n"
            "        try:\n"
            "            append_sheet_row(\n"
            "                tmp, 'Tasks',\n"
            "                {'Task': 'Task C', 'Status': 'In Progress', 'Points': 15},\n"
            "            )\n"
            "            wb2 = load_workbook(tmp)\n"
            "            ws2 = wb2['Tasks']\n"
            "            assert ws2.max_row == 4, \\\n"
            "                f'expected 4 rows total, got {ws2.max_row}'\n"
            "            passed += 1; print(f'\\u2705 Check 2: row count is now {ws2.max_row}')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 2: {e}')\n"
            "\n"
            "        # Check 3: new row values correct\n"
            "        try:\n"
            "            wb3 = load_workbook(tmp)\n"
            "            ws3 = wb3['Tasks']\n"
            "            new_row = list(ws3.iter_rows(\n"
            "                min_row=4, max_row=4, values_only=True\n"
            "            ))[0]\n"
            "            assert new_row[0] == 'Task C', \\\n"
            "                f'Task wrong: {new_row[0]!r}'\n"
            "            assert new_row[2] == 15, \\\n"
            "                f'Points wrong: {new_row[2]!r}'\n"
            "            passed += 1; print(f'\\u2705 Check 3: new row = {list(new_row)}')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 3: {e}')\n"
            "\n"
            "        # Check 4: existing rows unchanged\n"
            "        try:\n"
            "            wb4 = load_workbook(tmp)\n"
            "            ws4 = wb4['Tasks']\n"
            "            row2 = list(ws4.iter_rows(\n"
            "                min_row=2, max_row=2, values_only=True\n"
            "            ))[0]\n"
            "            assert row2[0] == 'Task A', \\\n"
            "                f'existing row 2 changed: {row2[0]!r}'\n"
            "            assert row2[2] == 10, \\\n"
            "                f'existing points changed: {row2[2]!r}'\n"
            "            passed += 1; print('\\u2705 Check 4: existing rows unchanged')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 4: {e}')\n"
            "\n"
            "        # Check 5: missing key appends None\n"
            "        try:\n"
            "            append_sheet_row(\n"
            "                tmp, 'Tasks',\n"
            "                {'Task': 'Task D'},  # Status and Points omitted\n"
            "            )\n"
            "            wb5 = load_workbook(tmp)\n"
            "            ws5 = wb5['Tasks']\n"
            "            row5 = list(ws5.iter_rows(\n"
            "                min_row=5, max_row=5, values_only=True\n"
            "            ))[0]\n"
            "            assert row5[0] == 'Task D', f'Task wrong: {row5[0]!r}'\n"
            "            assert row5[1] is None, \\\n"
            "                f'Status should be None for missing key, got {row5[1]!r}'\n"
            "            passed += 1; print('\\u2705 Check 5: missing key → None in row')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 5: {e}')\n"
            "\n"
            "    finally:\n"
            "        try:\n"
            "            os.unlink(tmp)\n"
            "        except Exception:\n"
            "            pass\n"
            "\n"
            "    if passed == total:\n"
            "        print('\\U0001f389 Exercise complete!')\n"
            "    print(f'\\nScore: {passed}/{total}')\n"
            "\n"
            "\n"
            "_run_checks()"
        ),
        md(
            "## Solution\n\n"
            "<details>\n"
            "<summary>Click to reveal</summary>\n\n"
            "```python\n"
            + APPEND_ROW_IMPL + "\n"
            "```\n\n"
            "</details>"
        ),
    ]


# ---------------------------------------------------------------------------
# Exercise 04 — bold_header_row
# ---------------------------------------------------------------------------

def ex04():
    global _cid; _cid = 300
    return [
        md(
            "# Day 028 — Exercise 4: bold_header_row\n\n"
            "**What you'll build:** `bold_header_row(path, sheet_name)` — "
            "opens an XLSX file, applies bold formatting to all cells in the first row, "
            "and saves the file.\n\n"
            "**Why it matters:** Bold headers are the single most impactful formatting "
            "change for a generated spreadsheet — they immediately communicate which row "
            "contains column names vs data. openpyxl's `Font(bold=True)` is the tool; "
            "`ws[1]` gives you the Cell objects (not values) needed to set `.font`."
        ),
        code(
            "from openpyxl import load_workbook, Workbook\n"
            "from openpyxl.styles import Font"
        ),
        md("## Your Implementation"),
        code(
            "def bold_header_row(path: str, sheet_name: str) -> None:\n"
            '    """\n'
            "    Apply bold formatting to the first row of a worksheet.\n\n"
            "    Args:\n"
            "        path:       Path to the .xlsx file (modified in place).\n"
            "        sheet_name: Name of the worksheet to format.\n"
            '    """\n'
            "    # TODO: wb = load_workbook(path). ws = wb[sheet_name]\n"
            "    # TODO: bold_font = Font(bold=True)\n"
            "    # TODO: for cell in ws[1]: cell.font = bold_font\n"
            "    # TODO: wb.save(path)\n"
            "    pass"
        ),
        md("## Check Your Work"),
        code(
            "import tempfile, os\n"
            "\n"
            "\n"
            "def _run_checks():\n"
            "    total = 5\n"
            "    passed = 0\n"
            "\n"
            "    tmp = tempfile.mktemp(suffix='.xlsx')\n"
            "    try:\n"
            "        # Setup: create test XLSX\n"
            "        wb = Workbook()\n"
            "        ws = wb.active\n"
            "        ws.title = 'Report'\n"
            "        ws.append(['Name', 'Value', 'Status'])\n"
            "        ws.append(['Alpha', 42, 'OK'])\n"
            "        ws.append(['Beta', 17, 'OK'])\n"
            "        wb.save(tmp)\n"
            "\n"
            "        # Check 1: defined\n"
            "        try:\n"
            "            assert 'bold_header_row' in globals()\n"
            "            passed += 1; print('\\u2705 Check 1: bold_header_row defined')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 1: {e}')\n"
            "            print(f'\\nScore: {passed}/{total}')\n"
            "            return\n"
            "\n"
            "        # Check 2: call does not raise\n"
            "        try:\n"
            "            bold_header_row(tmp, 'Report')\n"
            "            passed += 1; print('\\u2705 Check 2: bold_header_row ran without error')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 2: {e}')\n"
            "\n"
            "        # Check 3: first row cells are bold\n"
            "        try:\n"
            "            wb2 = load_workbook(tmp)\n"
            "            ws2 = wb2['Report']\n"
            "            for cell in ws2[1]:\n"
            "                assert cell.font.bold is True, \\\n"
            "                    f'cell {cell.coordinate} not bold: font.bold={cell.font.bold}'\n"
            "            passed += 1; print('\\u2705 Check 3: all header cells are bold')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 3: {e}')\n"
            "\n"
            "        # Check 4: data row cells are NOT bold (formatting not spread)\n"
            "        try:\n"
            "            wb3 = load_workbook(tmp)\n"
            "            ws3 = wb3['Report']\n"
            "            data_cell = ws3.cell(row=2, column=1)\n"
            "            assert data_cell.font.bold is not True, \\\n"
            "                f'data cell A2 should not be bold, got {data_cell.font.bold}'\n"
            "            passed += 1; print('\\u2705 Check 4: data row cells are not bold')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 4: {e}')\n"
            "\n"
            "        # Check 5: data is preserved (values unchanged)\n"
            "        try:\n"
            "            wb4 = load_workbook(tmp)\n"
            "            ws4 = wb4['Report']\n"
            "            header = [ws4.cell(row=1, column=c).value for c in range(1, 4)]\n"
            "            assert header == ['Name', 'Value', 'Status'], \\\n"
            "                f'header values changed: {header}'\n"
            "            row2_val = ws4.cell(row=2, column=1).value\n"
            "            assert row2_val == 'Alpha', \\\n"
            "                f'data changed: A2={row2_val!r}'\n"
            "            passed += 1; print('\\u2705 Check 5: cell values preserved after formatting')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 5: {e}')\n"
            "\n"
            "    finally:\n"
            "        try:\n"
            "            os.unlink(tmp)\n"
            "        except Exception:\n"
            "            pass\n"
            "\n"
            "    if passed == total:\n"
            "        print('\\U0001f389 Exercise complete!')\n"
            "    print(f'\\nScore: {passed}/{total}')\n"
            "\n"
            "\n"
            "_run_checks()"
        ),
        md(
            "## Solution\n\n"
            "<details>\n"
            "<summary>Click to reveal</summary>\n\n"
            "```python\n"
            + BOLD_HEADER_IMPL + "\n"
            "```\n\n"
            "</details>"
        ),
    ]


# ---------------------------------------------------------------------------
# Exercise 05 — ai_analyze_sheet
# ---------------------------------------------------------------------------

def ex05():
    global _cid; _cid = 400
    return [
        md(
            "# Day 028 — Exercise 5: ai_analyze_sheet\n\n"
            "**What you'll build:** `ai_analyze_sheet(path, question, model='llama3.2') -> str` — "
            "reads rows from an XLSX file, converts them to JSON for the LLM context, "
            "and returns an AI-generated answer to the given question.\n\n"
            "**Why it matters:** This is the AI layer of today's pipeline — the same pattern "
            "as `ai_generate_section` (Day 26): system prompt establishes persona, user prompt "
            "provides data and question, function returns content string. The `rows[:20]` cap "
            "keeps the data within the LLM's context window even for large sheets."
        ),
        code(
            "import json\n"
            "import ollama\n"
            "from openpyxl import load_workbook, Workbook"
        ),
        md("## Provided: read_sheet_rows"),
        code(READ_ROWS_IMPL),
        md("## Your Implementation"),
        code(
            "def ai_analyze_sheet(\n"
            "    path: str,\n"
            "    question: str,\n"
            "    model: str = 'llama3.2',\n"
            ") -> str:\n"
            '    """\n'
            "    Read an XLSX file and ask an LLM a question about its data.\n\n"
            "    Args:\n"
            "        path:     Path to the .xlsx file.\n"
            "        question: Question to ask about the data.\n"
            "        model:    Ollama model name.\n\n"
            "    Returns:\n"
            "        AI-generated answer string.\n"
            '    """\n'
            "    # TODO: rows = read_sheet_rows(path)\n"
            "    # TODO: data_str = json.dumps(rows[:20], indent=2)\n"
            "    # TODO: system: 'data analyst; answer concisely'\n"
            "    # TODO: user: f'Data:\\n{data_str[:3000]}\\n\\nQuestion: {question}'\n"
            "    # TODO: return response['message']['content']\n"
            "    pass"
        ),
        md("## Check Your Work"),
        code(
            "import tempfile, os\n"
            "\n"
            "\n"
            "def _run_checks():\n"
            "    total = 5\n"
            "    passed = 0\n"
            "\n"
            "    tmp = tempfile.mktemp(suffix='.xlsx')\n"
            "    try:\n"
            "        # Setup: create XLSX with sample data\n"
            "        wb = Workbook()\n"
            "        ws = wb.active\n"
            "        ws.title = 'Sales'\n"
            "        ws.append(['Product', 'Q1', 'Q2', 'Q3'])\n"
            "        ws.append(['Alpha', 1200, 1350, 980])\n"
            "        ws.append(['Beta',   870,  920, 1100])\n"
            "        ws.append(['Gamma', 2100, 1980, 2200])\n"
            "        wb.save(tmp)\n"
            "\n"
            "        # Check 1: defined\n"
            "        try:\n"
            "            assert 'ai_analyze_sheet' in globals()\n"
            "            passed += 1; print('\\u2705 Check 1: ai_analyze_sheet defined')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 1: {e}')\n"
            "            print(f'\\nScore: {passed}/{total}')\n"
            "            return\n"
            "\n"
            "        result = None\n"
            "\n"
            "        # Check 2: returns a string\n"
            "        try:\n"
            "            result = ai_analyze_sheet(\n"
            "                tmp, 'Which product has the highest Q1 sales?'\n"
            "            )\n"
            "            assert isinstance(result, str), \\\n"
            "                f'expected str, got {type(result)}'\n"
            "            passed += 1; print('\\u2705 Check 2: returns a string')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 2: {e}')\n"
            "\n"
            "        # Check 3: result is non-empty\n"
            "        try:\n"
            "            assert result is not None\n"
            "            assert len(result.strip()) > 10, \\\n"
            "                f'response too short ({len(result)} chars): {result!r}'\n"
            "            passed += 1; print(f'\\u2705 Check 3: response is {len(result)} chars')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 3: {e}')\n"
            "\n"
            "        # Check 4: different question gives a response\n"
            "        try:\n"
            "            result2 = ai_analyze_sheet(\n"
            "                tmp, 'How many products are in the data?'\n"
            "            )\n"
            "            assert isinstance(result2, str) and len(result2) > 5\n"
            "            passed += 1; print('\\u2705 Check 4: different question also works')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 4: {e}')\n"
            "\n"
            "        # Check 5: result references relevant data (product name present)\n"
            "        try:\n"
            "            assert result is not None\n"
            "            # Gamma has highest Q1 (2100) — LLM should mention it\n"
            "            assert 'Gamma' in result or 'gamma' in result.lower(), \\\n"
            "                f'expected Gamma mentioned in Q1-highest answer: {result[:200]}'\n"
            "            passed += 1; print('\\u2705 Check 5: answer references the correct product')\n"
            "        except Exception as e:\n"
            "            print(f'\\u274c Check 5: {e}')\n"
            "\n"
            "    finally:\n"
            "        try:\n"
            "            os.unlink(tmp)\n"
            "        except Exception:\n"
            "            pass\n"
            "\n"
            "    if passed == total:\n"
            "        print('\\U0001f389 Exercise complete!')\n"
            "    print(f'\\nScore: {passed}/{total}')\n"
            "\n"
            "\n"
            "_run_checks()"
        ),
        md(
            "## Solution\n\n"
            "<details>\n"
            "<summary>Click to reveal</summary>\n\n"
            "```python\n"
            + AI_ANALYZE_IMPL + "\n"
            "```\n\n"
            "</details>"
        ),
    ]


# ---------------------------------------------------------------------------
# Project notebook (student template — NOT executed by gate)
# ---------------------------------------------------------------------------

def project_nb():
    global _cid; _cid = 500
    return [
        md(
            "# Day 028 Project: SpreadsheetAnalyst\n\n"
            "## What You're Building\n\n"
            "A `SpreadsheetAnalyst` class that:\n"
            "1. Reads an XLSX file using `read_sheet_rows`\n"
            "2. Analyzes the data with an LLM (using `analyze`)\n"
            "3. Writes a formatted two-sheet output with `write_summary`:\n"
            "   - **Data** sheet — source rows with bold header\n"
            "   - **Summary** sheet — AI analysis text\n\n"
            "## Project Requirements\n\n"
            "1. Implement `SpreadsheetAnalyst` with:\n"
            "   - `analyze(rows, question) -> str` — AI analysis\n"
            "   - `write_summary(input_path, analysis, output_path) -> None` — two-sheet report\n"
            "   - `run(input_path, question, output_path) -> dict` — full pipeline\n"
            "2. Create sample data, write it to an XLSX with `write_sheet_rows`\n"
            "3. Call `analyst.run(...)` and store as `result`\n"
            "4. Verify with `_run_project_checks()`"
        ),
        code(
            "import json, os, tempfile\n"
            "from openpyxl import load_workbook, Workbook\n"
            "from openpyxl.styles import Font\n"
            "import ollama"
        ),
        md("## Provided: Helper Functions"),
        code(ALL_IMPLS),
        md(
            "## Your Implementation\n\n"
            "Implement `SpreadsheetAnalyst` using the helper functions above."
        ),
        code(
            "class SpreadsheetAnalyst:\n"
            "    def __init__(self, model: str = 'llama3.2'):\n"
            "        self.model = model\n"
            "\n"
            "    def analyze(self, rows: list[dict], question: str) -> str:\n"
            "        # TODO: data_str = json.dumps(rows[:20], indent=2)\n"
            "        # TODO: ollama.chat with data analyst system prompt\n"
            "        # TODO: user: f'Data:\\n{data_str[:3000]}\\n\\nQuestion: {question}'\n"
            "        # TODO: return response['message']['content']\n"
            "        pass\n"
            "\n"
            "    def write_summary(\n"
            "        self, input_path: str, analysis: str, output_path: str\n"
            "    ) -> None:\n"
            "        # TODO: rows = read_sheet_rows(input_path)\n"
            "        # TODO: wb = Workbook(). ws_data = wb.active. ws_data.title = 'Data'\n"
            "        # TODO: write headers + rows + bold header to ws_data\n"
            "        # TODO: ws_sum = wb.create_sheet('Summary')\n"
            "        # TODO: ws_sum.append(['AI Analysis']). ws_sum.append([analysis])\n"
            "        # TODO: wb.save(output_path)\n"
            "        pass\n"
            "\n"
            "    def run(\n"
            "        self, input_path: str, question: str, output_path: str\n"
            "    ) -> dict:\n"
            "        # TODO: rows = read_sheet_rows(input_path)\n"
            "        # TODO: analysis = self.analyze(rows, question)\n"
            "        # TODO: self.write_summary(input_path, analysis, output_path)\n"
            "        # TODO: return {'rows_analyzed': len(rows),\n"
            "        #              'analysis': analysis, 'output': output_path}\n"
            "        pass"
        ),
        md("## Sample Data and Run"),
        code(
            "SAMPLE_DATA = [\n"
            "    {'Product': 'Alpha Pro', 'Q1': 1200, 'Q2': 1350, 'Q3': 980,  'Q4': 1600},\n"
            "    {'Product': 'Beta Max',  'Q1':  870, 'Q2':  920, 'Q3': 1100, 'Q4': 1250},\n"
            "    {'Product': 'Gamma Plus','Q1': 2100, 'Q2': 1980, 'Q3': 2200, 'Q4': 2450},\n"
            "    {'Product': 'Delta Lite','Q1':  450, 'Q2':  510, 'Q3':  390, 'Q4':  620},\n"
            "    {'Product': 'Epsilon X', 'Q1': 3200, 'Q2': 3100, 'Q3': 3450, 'Q4': 3600},\n"
            "]\n"
            "HEADERS = ['Product', 'Q1', 'Q2', 'Q3', 'Q4']\n"
            "QUESTION = 'Which product has the highest total sales? What growth trends do you see?'"
        ),
        code(
            "# Step 1: write sample data to XLSX\n"
            "# input_path = os.path.join(tempfile.gettempdir(), 'day028_input.xlsx')\n"
            "# write_sheet_rows(input_path, 'Sales', HEADERS, SAMPLE_DATA)\n"
            "\n"
            "# Step 2: run the analyst\n"
            "# output_path = os.path.join(tempfile.gettempdir(), 'day028_report.xlsx')\n"
            "# analyst = SpreadsheetAnalyst()\n"
            "# result = analyst.run(input_path, QUESTION, output_path)\n"
            "# print(f\"Rows analyzed: {result['rows_analyzed']}\")\n"
            "# print(f\"Output: {result['output']}\")\n"
            "# print(f\"Analysis preview: {result['analysis'][:200]}\")"
        ),
        md("## Checks"),
        code(
            "def _run_project_checks():\n"
            "    total = 5\n"
            "    passed = 0\n"
            "\n"
            "    # Check 1: SpreadsheetAnalyst has all required methods\n"
            "    try:\n"
            "        assert 'SpreadsheetAnalyst' in globals()\n"
            "        for m in ('analyze', 'write_summary', 'run'):\n"
            "            assert hasattr(SpreadsheetAnalyst, m), \\\n"
            "                f'SpreadsheetAnalyst missing method: {m}'\n"
            "        passed += 1; print('\\u2705 Check 1: all methods present')\n"
            "    except Exception as e:\n"
            "        print(f'\\u274c Check 1: {e}')\n"
            "\n"
            "    # Check 2: analyst is an instance\n"
            "    try:\n"
            "        assert 'analyst' in globals()\n"
            "        assert isinstance(analyst, SpreadsheetAnalyst)\n"
            "        passed += 1; print('\\u2705 Check 2: analyst is a SpreadsheetAnalyst')\n"
            "    except Exception as e:\n"
            "        print(f'\\u274c Check 2: {e}')\n"
            "\n"
            "    # Check 3: result dict has required keys\n"
            "    try:\n"
            "        assert 'result' in globals()\n"
            "        for k in ('rows_analyzed', 'analysis', 'output'):\n"
            "            assert k in result, f\"result missing '{k}': {list(result)}\"\n"
            "        passed += 1; print('\\u2705 Check 3: result has rows_analyzed/analysis/output')\n"
            "    except Exception as e:\n"
            "        print(f'\\u274c Check 3: {e}')\n"
            "\n"
            "    # Check 4: output file exists with two sheets\n"
            "    try:\n"
            "        assert 'result' in globals()\n"
            "        out = result.get('output', '')\n"
            "        assert os.path.exists(out), f'output file not found: {out}'\n"
            "        wb = load_workbook(out)\n"
            "        assert 'Data' in wb.sheetnames, \\\n"
            "            f\"'Data' sheet missing: {wb.sheetnames}\"\n"
            "        assert 'Summary' in wb.sheetnames, \\\n"
            "            f\"'Summary' sheet missing: {wb.sheetnames}\"\n"
            "        passed += 1; print(f'\\u2705 Check 4: output has Data + Summary sheets')\n"
            "    except Exception as e:\n"
            "        print(f'\\u274c Check 4: {e}')\n"
            "\n"
            "    # Check 5: analysis is a non-empty string\n"
            "    try:\n"
            "        assert 'result' in globals()\n"
            "        analysis = result.get('analysis', '')\n"
            "        assert isinstance(analysis, str) and len(analysis) > 20, \\\n"
            "            f'analysis should be non-empty str: {analysis!r}'\n"
            "        passed += 1; print(f'\\u2705 Check 5: analysis is {len(analysis)} chars')\n"
            "    except Exception as e:\n"
            "        print(f'\\u274c Check 5: {e}')\n"
            "\n"
            "    if passed == total:\n"
            "        print('\\U0001f389 Project complete!')\n"
            "    print(f'\\nScore: {passed}/{total}')\n"
            "\n"
            "\n"
            "_run_project_checks()"
        ),
        md(
            "## Bonus Challenges\n\n"
            "- Add `autofit_columns(path, sheet_name)` and apply it to the Data sheet\n"
            "- Add a `charts` argument to `write_summary` — use openpyxl's `BarChart` to "
            "embed a bar chart of the data in the Data sheet\n"
            "- Add a `to_google_sheets(spreadsheet_key, creds_path)` method that uploads "
            "the Data sheet rows to a Google Sheet using gspread\n"
            "- Add a `summarize_by_column(path, col_name, model)` function that groups "
            "rows by a column's unique values and generates an AI summary for each group\n"
            "- Handle large sheets: if `len(rows) > 100`, sample 20 rows at random "
            "for the AI context instead of taking the first 20"
        ),
    ]


# ---------------------------------------------------------------------------
# Solution notebook (executed by gate)
# ---------------------------------------------------------------------------

def solution_nb():
    global _cid; _cid = 600

    imports = (
        "import json, os, tempfile\n"
        "from openpyxl import load_workbook, Workbook\n"
        "from openpyxl.styles import Font\n"
        "import ollama"
    )

    all_code = imports + "\n\n\n" + ALL_IMPLS + "\n\n\n" + ANALYST_IMPL

    sample_data = (
        "SAMPLE_DATA = [\n"
        "    {'Product': 'Alpha Pro',  'Q1': 1200, 'Q2': 1350, 'Q3':  980, 'Q4': 1600},\n"
        "    {'Product': 'Beta Max',   'Q1':  870, 'Q2':  920, 'Q3': 1100, 'Q4': 1250},\n"
        "    {'Product': 'Gamma Plus', 'Q1': 2100, 'Q2': 1980, 'Q3': 2200, 'Q4': 2450},\n"
        "    {'Product': 'Delta Lite', 'Q1':  450, 'Q2':  510, 'Q3':  390, 'Q4':  620},\n"
        "    {'Product': 'Epsilon X',  'Q1': 3200, 'Q2': 3100, 'Q3': 3450, 'Q4': 3600},\n"
        "]\n"
        "HEADERS = ['Product', 'Q1', 'Q2', 'Q3', 'Q4']"
    )

    return [
        md(
            "# Day 028 Project Solution — SpreadsheetAnalyst\n\n"
            "A `SpreadsheetAnalyst` that reads tabular data, generates AI insights, "
            "and writes a formatted two-sheet report."
        ),
        code(all_code),
        md("## Action 1 — Create Sample XLSX with Bold Headers"),
        code(
            sample_data + "\n"
            "\n"
            "input_path = os.path.join(tempfile.gettempdir(), 'day028_input.xlsx')\n"
            "write_sheet_rows(input_path, 'Sales', HEADERS, SAMPLE_DATA)\n"
            "bold_header_row(input_path, 'Sales')\n"
            "\n"
            "rows = read_sheet_rows(input_path, 'Sales')\n"
            "print(f'Created input XLSX with {len(rows)} data rows')\n"
            "print(f'Columns: {list(rows[0].keys())}')\n"
            "print(f'First row: {rows[0]}')"
        ),
        md("## Action 2 — Analyze Data with AI"),
        code(
            "QUESTION = (\n"
            "    'Which product has the highest total annual sales? '\n"
            "    'What growth trends do you observe across the quarters?'\n"
            ")\n"
            "\n"
            "analyst = SpreadsheetAnalyst()\n"
            "analysis = analyst.analyze(rows, QUESTION)\n"
            "print('AI Analysis:')\n"
            "print(analysis)"
        ),
        md("## Action 3 — Write Formatted Report and Verify"),
        code(
            "output_path = os.path.join(tempfile.gettempdir(), 'day028_report.xlsx')\n"
            "analyst.write_summary(input_path, analysis, output_path)\n"
            "\n"
            "# Verify the output\n"
            "wb = load_workbook(output_path)\n"
            "print(f'Output sheets: {wb.sheetnames}')\n"
            "\n"
            "ws_data = wb['Data']\n"
            "print(f'Data sheet rows: {ws_data.max_row} (including header)')\n"
            "print(f'Header row bold: {ws_data[\"A1\"].font.bold}')\n"
            "\n"
            "ws_sum = wb['Summary']\n"
            "summary_cell = ws_sum.cell(row=2, column=1).value or ''\n"
            "print(f'Summary sheet preview: {summary_cell[:100]}')\n"
            "\n"
            "print('\\nAnalysis complete!')"
        ),
    ]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Generating Day 028 notebooks...")
    ex_dir = DAY_DIR / "exercises"
    proj_dir = DAY_DIR / "project"
    sol_dir = proj_dir / "solution"
    for d in (ex_dir, proj_dir, sol_dir):
        d.mkdir(parents=True, exist_ok=True)

    write_nb(ex_dir / "exercise_01.ipynb", ex01())
    write_nb(ex_dir / "exercise_02.ipynb", ex02())
    write_nb(ex_dir / "exercise_03.ipynb", ex03())
    write_nb(ex_dir / "exercise_04.ipynb", ex04())
    write_nb(ex_dir / "exercise_05.ipynb", ex05())
    write_nb(proj_dir / "project.ipynb", project_nb())
    write_nb(sol_dir / "solution.ipynb", solution_nb())
    print("Done.")


if __name__ == "__main__":
    main()
